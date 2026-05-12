import calendar
from django.http import HttpResponse, HttpResponseForbidden
from django.http import HttpResponseBadRequest
from django.shortcuts import get_object_or_404
from django.views.decorators.http import require_http_methods
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side 
from .models import Lab, LabMembership, User, WorkEntry, MonthlyMeta
from .excel_utils import apply_border
import zipfile
from io import BytesIO
from collections import defaultdict
from openpyxl.comments import Comment
from openpyxl.utils import range_boundaries
from openpyxl.drawing.image import Image as XLImage

def is_merged_cell(ws, row, col):
    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        if min_row <= row <= max_row and min_col <= col <= max_col:
            return not (row == min_row and col == min_col)
    return False

def merge_center(ws, start_row, start_col, end_row, end_col, value=None, fill=None, font=None):
    ws.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)
    cell = ws.cell(start_row, start_col)
    if value is not None:
        cell.value = value
    if fill is not None:
        cell.fill = fill
    if font is not None:
        cell.font = font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def months_to_RO(month):
    lunile = ["ianuarie","februarie", "martie", "aprile",
              "mai", "iunie","iulie", "august",
              "septembrie", "octombrie", "noiembrie", "decembrie"]
    return lunile[month-1]

def get_initials(user):
    first = user.first_name[:1].upper() if user.first_name else ""
    last = user.last_name[:1].upper() if user.last_name else ""
    return f"{first}{last}"


def autofit_sheet(ws):
    """Adjust column widths and row heights based on cell content.
    Skips merged cells except for the top-left cell of each merge range.
    """
    merged_starts = {(r.min_row, r.min_col): r for r in ws.merged_cells.ranges}

    min_width = 8
    max_width = 55
    padding = 2

    wrap_columns = set()
    computed_widths = {}

    # columns (approximate "autofit" + clamp so long text doesn't explode layout)
    for col_cells in ws.columns:
        max_len = 0
        has_newlines = False
        column = col_cells[0].column_letter

        for cell in col_cells:
            if cell.value is None:
                continue

            merged = merged_starts.get((cell.row, cell.column))
            # only measure the top-left cell of a merge
            if merged and (cell.row != merged.min_row or cell.column != merged.min_col):
                continue

            text = str(cell.value)
            lines = text.splitlines()
            has_newlines = has_newlines or (len(lines) > 1)
            if not lines:
                continue

            raw_len = max((len(l) for l in lines))
            # if the cell spans multiple cols, distribute width
            colspan = 1
            if merged:
                colspan = merged.max_col - merged.min_col + 1
            effective_len = (raw_len + colspan - 1) // colspan  # ceil division
            max_len = max(max_len, effective_len)

        width = max_len + padding
        if width > max_width:
            wrap_columns.add(column)
        if has_newlines:
            wrap_columns.add(column)

        width = max(min_width, min(max_width, width))
        computed_widths[column] = width
        ws.column_dimensions[column].width = width

    # Enable wrapping for columns that contain long/multiline values.
    if wrap_columns:
        for column in wrap_columns:
            for cell in ws[column]:
                if cell.value is None:
                    continue
                merged = merged_starts.get((cell.row, cell.column))
                if merged and (cell.row != merged.min_row or cell.column != merged.min_col):
                    continue
                cell.alignment = cell.alignment.copy(wrap_text=True)

    # rows
    for row_cells in ws.iter_rows():
        max_lines = 1
        for cell in row_cells:
            if cell.value is None:
                continue
            merged = merged_starts.get((cell.row, cell.column))
            if merged and (cell.row != merged.min_row or cell.column != merged.min_col):
                continue
            text = str(cell.value)
            lines = text.splitlines() or [""]

            line_count = len(lines)
            column_letter = cell.column_letter
            if column_letter in wrap_columns and line_count == 1:
                # Roughly estimate wrapped lines from current column width.
                merged_width = computed_widths.get(column_letter, max_width)
                if merged:
                    # Sum widths across spanned columns (approx).
                    merged_width = 0
                    for col in range(merged.min_col, merged.max_col + 1):
                        letter = ws.cell(row=1, column=col).column_letter
                        merged_width += computed_widths.get(letter, max_width)

                chars_per_line = max(1, int(merged_width))
                est = (len(lines[0]) + chars_per_line - 1) // chars_per_line
                line_count = max(line_count, est)

            max_lines = max(max_lines, line_count)

        # Clamp to avoid extreme heights from very long text.
        ws.row_dimensions[row_cells[0].row].height = min(220, max_lines * 16)


def upt_workbook(lab, users, month, year, director):
    days_in_month = calendar.monthrange(year, month)[1]
    month_name = months_to_RO(month)

    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    bold = Font(bold=True)
    center_wrapped = Alignment(horizontal="center", vertical="center", wrap_text=True)

    weekend_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="lightHorizontal")

    last_col = days_in_month + 2

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Create a sheet for each user
    for user in users:
        ws = wb.create_sheet(title=f"{user.first_name}{user.last_name}")
        activitati = list(lab.activitati.all().order_by("id"))
        activitate_index = {act.id: i for i, act in enumerate(activitati)}

        # HEADER
        ws["A1"] = "Universitatea Politehnica Timișoara"
        ws["A2"] = f"IPCEI - Laboratorul L{lab.name[-1]}"

        ws["A7"] = "FOAIE INDIVIDUALĂ DE PREZENȚĂ - EVIDENȚA NUMĂRULUI DE ORE LUCRATE (PONTAJ)"
        ws.merge_cells(start_row=7, start_column=1, end_row=7, end_column=last_col)
        ws.cell(row=7, column=1).alignment = center

        ws["A8"] = "Titlul proiectului: \"Investiția 14. Proiecte transfrontaliere și multinaționale – Procesoare cu consum redus de energie și cipuri semiconductoare \", Componenta C9. Suport pentru sectorul privat, cercetare, dezvoltare și inovare,\n \
              Pilonul III. Creștere inteligentă, sustenabilă și favorabilă incluziunii, inclusiv coeziune economică, locuri de muncă, productivitate, competitivitate, cercetare, dezvoltare și inovare, precum și o piață internă funcțională,\n" \
              " cu întreprinderi mici și mijlocii (IMM-uri) puternice – Subproiectul cu titlul „ Traductoare inteligente eficiente din punct de vedere energetic pentru sisteme auto – inovație de-a lungul lanțului valoric (ASSET-IxC –UPT)\" "
        ws.merge_cells(start_row=8, start_column=1, end_row=8, end_column=last_col)
        ws.cell(row=8, column=1).alignment = center

        ws["A9"] = "Nr. Ordine 5.PI/I4/C9"
        ws.merge_cells(start_row=9, start_column=1, end_row=9, end_column=last_col)
        ws.cell(row=9, column=1).alignment = center

        ws["A10"] = f"{month_name} {year}"
        ws.merge_cells(start_row=10, start_column=1, end_row=10, end_column=last_col)
        ws.cell(row=10, column=1).font = bold
        ws.cell(row=10, column=1).alignment = center

        row = 12

        ws.cell(row=row, column=1, value="Nume și prenume cadru didactic/salariat").font = bold
        ws.merge_cells(start_row=row, start_column=1, end_row=row+1, end_column=1)
        ws.cell(row=row, column=1).alignment = center_wrapped

        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=last_col-1)

        ws.cell(row=row+2, column=1, value=f"{user.last_name.upper()} {user.first_name}").font = bold
        ws.merge_cells(start_row=row+2, start_column=1, end_row=row+5, end_column=1)
        ws.cell(row=row+2, column=1).alignment = center_wrapped

        ws.cell(row=row+6, column=1, value="Total ore lucrate:").font = bold
        ws.cell(row=row+6, column=1).alignment = center_wrapped

        # DAYS HEADER
        for d in range(1, days_in_month + 1):
            weekday = calendar.weekday(year, month, d)
            col = d + 1

            cell = ws.cell(row=row+1, column=col, value=f"{d}\n{month_name[:3]}.")
            cell.alignment = center
            cell.font = bold

            if weekday >= 5:
                for r in range(row+2, row+6):
                    ws.cell(row=r, column=col).fill = weekend_fill

        ws.cell(row=row, column=days_in_month+2, value="Semnătură")
        ws.merge_cells(start_row=row, start_column=days_in_month+2, end_row=row+1, end_column=days_in_month+2)
        ws.cell(row=row, column=days_in_month+2).alignment = center

        ws.merge_cells(start_row=row+2, start_column=days_in_month+2, end_row=row+5, end_column=days_in_month+2)

        membership = LabMembership.objects.filter(lab=lab, profile__user=user).first()
        if membership and membership.signature_png:
            try:
                img_stream = BytesIO(membership.signature_png)
                img = XLImage(img_stream)

                max_width = 160
                max_height = 100
                scale = min(max_width / float(img.width or 1), max_height / float(img.height or 1), 1.0)
                img.width = int((img.width or 1) * scale)
                img.height = int((img.height or 1) * scale)

                anchor_col_letter = ws.cell(row=1, column=days_in_month + 2).column_letter
                img.anchor = f"{anchor_col_letter}{row+2}"
                ws.add_image(img)
            except Exception:
                ws.cell(row=row+2, column=days_in_month + 2, value="Semnatura PNG invalida")

        # FETCH ENTRIES
        entries = WorkEntry.objects.filter(
            user=user,
            lab=lab,
            date__year=year,
            date__month=month
        )

        hours = defaultdict(float)
        durata = defaultdict(list)

        for e in entries:
            key = (e.activitate_id, e.date.day)
            hours[key] += e.nr_ore
            durata[key].append(str(e.durata))

        total = 0
        for d in range(1, days_in_month + 1):
            daily_hours_sum = 0
            intervaluri_durata = []

            # Sort activities by index to maintain deterministic order
            sorted_acts = sorted(activitate_index.items(), key=lambda x: x[1])

            for act_id, i in sorted_acts:
                h = hours.get((act_id, d), 0)
                if isinstance(h, (int, float)):
                    daily_hours_sum += h
                dur_list = durata.get((act_id, d), [])
                intervaluri_durata.extend(dur_list)

            intervaluri_durata.sort(key=lambda x: int(x.split('-')[0].replace(':', '')) if '-' in x and ':' in x.split('-')[0] else 0)

            for idx, dur_val in enumerate(intervaluri_durata):
                row_offset = min(idx, 3)  # Use max 4 rows (offsets 0-3)
                target_row = row + 2 + row_offset
                current_val = ws.cell(row=target_row, column=d + 1).value
                new_val = f"{current_val}\n{dur_val}" if (idx >= 4 and current_val) else dur_val
                ws.cell(row=target_row, column=d + 1, value=new_val).alignment = center

            ws.cell(row=row + 6, column=d + 1, value=daily_hours_sum).font = bold
            total += daily_hours_sum

        ws.cell(row=row + 6, column=days_in_month + 2, value=total).font = bold

        for r in range(row, row+7):
            for c in range(1, last_col+1):
                ws.cell(row=r, column=c).border = thin_border

        # auto‑size sheet
        autofit_sheet(ws)

        final_row = row + 8
        ws.cell(row=final_row, column=1, value="Manager proiect").alignment = center_wrapped
        ws.cell(row=final_row+1, column=1, value="Conf.univ.dr.ing. Ionel Raul-Ciprian").alignment = center_wrapped

        ws.cell(row=final_row, column=last_col-2, value="Întocmit,")
        ws.cell(row=final_row+1, column=last_col-2, value=f"{user.last_name} {user.first_name}")
        if membership and membership.signature_png:
            try:
                img_stream = BytesIO(membership.signature_png)
                img = XLImage(img_stream)

                max_width = 160
                max_height = 100
                scale = min(max_width / float(img.width or 1), max_height / float(img.height or 1), 1.0)
                img.width = int((img.width or 1) * scale)
                img.height = int((img.height or 1) * scale)

                anchor_col_letter = ws.cell(row=1, column=last_col-2).column_letter
                img.anchor = f"{anchor_col_letter}{final_row+2}"
                ws.add_image(img)
            except Exception:
                ws.cell(row=final_row+2, column=last_col-2, value="Semnatura PNG invalida")

    return wb


def normalize_url(value):
        value = str(value or "").strip()
        if not value:
            return ""
        if value.startswith(("http://", "https://", "mailto:")):
            return value
        if value.startswith("www."):
            return f"https://{value}"
        return ""

def write_hyperlink(ws, row, col, raw_value):
    link_font = Font(color="215C98", underline="single")
    text = str(raw_value or "").strip()
    cell = ws.cell(row=row, column=col, value=text)
    url = normalize_url(text)
    if url:
        cell.hyperlink = url
        cell.font = link_font
    return cell

def conti_workbook(lab, users, month, year, director):

    days = calendar.monthrange(year, month)[1]
    month_name = months_to_RO(month)

    user_workbooks = []

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    center_wrapped = Alignment(horizontal="center", vertical="center", wrap_text=True)
    bold = Font(bold=True)
    color = Font(color="C00000", bold=True)
    wb_fill = PatternFill(start_color="DAF2D0", end_color="DAF2D0", fill_type="solid") 

    activitati = list(lab.activitati.all().order_by("id"))
    activitate_index = {act.id: i for i, act in enumerate(activitati)}
    activitate_ids = {act.id for act in activitati}
    nr_Activitati = len(activitati)

    membership_by_user_id = {
        m.profile.user_id: m
        for m in LabMembership.objects.filter(lab=lab, profile__user__in=users).select_related(
            "profile__user"
        )
    }

    for user in users:

        wb = Workbook()
        ws = wb.active
        ws.title = f"{user.last_name}_{user.first_name}"

        # ---------------- HEADER ----------------
        ws["A1"] = "Info:"
        ws["B1"] = f"Activitati de cercetare fundamentala in cadrul {lab}"
        ws["A2"] = (lab.titlu or "").strip()
        ws.cell(row=2, column=1).font = color
        ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=2)
        ws["A3"] = "PI:"
        ws["B3"] = "Universitatea Politehnica Timisoara"
        ws["A4"] = "Traductoare inteligente eficiente din punct de vedere energetic pentru sisteme auto  – inovație de-a lungul lanțului valoric (ASSET-IxC - UPT) "
        ws.merge_cells(start_row=4,start_column=1,end_row=4,end_column=2)
        ws.cell(row=4, column=1).font = color
        ws["A5"] = "NR:"
        ws["B5"] = "5.PI/I4/C9"

        ws["A7"] = "PONTAJ INDIVIDUAL PENTRU LUNA/ANUL:"
        ws["B7"] = f"{month_name.upper()}/{year}"
        ws.cell(row=7, column=1).font = bold

        ws["A8"] = "NUME SI PRENUME:"
        ws["B8"] = f"{user.last_name.upper()} {user.first_name}"
        ws.cell(row=8, column=1).font = bold

        ws["A10"] = "POST PROIECT"
        membership = membership_by_user_id.get(user.id)
        ws["B10"] = (membership.post or "") if membership else ""
        ws.cell(row=10, column=1).font = bold

        for r in (7,8,10):
            ws.cell(r,2).font = color
        for r in (1,3,5):
            for c in (1,2):
                ws.cell(r,c).font = bold
        for r in range(1,11):
            for c in range(1,3):
                ws.cell(r,c).alignment = center

        # Thin borders for the small header blocks
        apply_border(ws, 1, 1, 5, 2, thickness="thin")   # A1:B5
        apply_border(ws, 7, 1, 8, 2, thickness="thin")   # A7:B8
        apply_border(ws, 10, 1, 10, 2, thickness="thin") # A10

        # ------------- ACTIVITIES TABLE -------------
        start = 12

        ws.cell(start,1,"Activitate")
        ws.cell(start,2,"Denumire activitate sumar")
        ws.cell(start,3,"Scurta descriere activitate")
        ws.cell(start,4,"Link")
        ws.cell(start,5,"Livrabile")

        ws.cell(start,6,"Mapare activitate partener direct")
        ws.cell(start,7,"Comentarii/Scurta descriere actiuni cercetare")
        for c in range(1,8):
            ws.cell(start,c).font = bold
            ws.cell(start,c).alignment = center_wrapped

        entries = (
            WorkEntry.objects.filter(
                user=user,
                lab=lab,
                date__year=year,
                date__month=month,
            )
            .select_related("activitate")
            .order_by("date", "id")
        )

        # Monthly fields (links/livrabile/comentarii) are now user-entered and saved
        # separately from daily entries.
        links_by_act = {}
        livrabile_by_act = {}
        comentarii_by_act = {}

        metas = MonthlyMeta.objects.filter(
            user=user,
            lab=lab,
            month=month,
            year=year,
            activitate_id__in=list(activitate_ids),
        )
        for m in metas:
            if m.links:
                links_by_act[m.activitate_id] = str(m.links).strip()
            if m.livrabil:
                livrabile_by_act[m.activitate_id] = str(m.livrabil).strip()
            if m.comentarii:
                comentarii_by_act[m.activitate_id] = str(m.comentarii).strip()

        # Backwards-compatible fallback: if no MonthlyMeta exists for an activity yet,
        # take the latest non-empty values from that user's daily entries.
        for e in entries:
            act_id = e.activitate_id
            if act_id not in activitate_ids:
                continue
            if act_id not in links_by_act and e.links:
                links_by_act[act_id] = str(e.links).strip()
            if act_id not in livrabile_by_act and e.livrabil:
                livrabile_by_act[act_id] = str(e.livrabil).strip()
            if act_id not in comentarii_by_act and e.comentarii:
                comentarii_by_act[act_id] = str(e.comentarii).strip()

        # Fill the activity description block from admin-maintained Activitate.descriere.
        for i, act in enumerate(activitati):
            r = start + 1 + i
            ws.cell(r, 2, (act.denumire_activitate).strip())
            ws.cell(row=r, column=2).alignment = center_wrapped

            ws.cell(r, 3, act.descriere)
            ws.cell(row=r, column=3).alignment = center_wrapped
            act_id = act.id
            # Column 4: Link (from WorkEntry.links)
            link_cell = write_hyperlink(ws, r, 4, links_by_act.get(act_id, ""))
            link_cell.alignment = center_wrapped

            # Column 5: Livrabile (from WorkEntry.livrabil)
            liv_cell = write_hyperlink(ws, r, 5, livrabile_by_act.get(act_id, ""))
            liv_cell.alignment = center_wrapped

            # Comentarii (from WorkEntry.comentarii)
            comm_cell = ws.cell(r, 7, comentarii_by_act.get(act_id, ""))
            comm_cell.alignment = center_wrapped

        for i in range(1,9):    
            r = start+i
            ws.cell(r,1,f"A{i}").font = bold
            ws.cell(row=r, column=1).alignment = center_wrapped
            if ws.cell(row=r, column=2).value is None:
                ws.cell(row=r, column=2, value = "NU ESTE CAZUL").alignment = center_wrapped
                ws.cell(row=r, column=2, value = "NU ESTE CAZUL").font = color



        # Thin borders for the activitati block
        activitati_end_row = start + 8
        last_activitati_col = 7
        apply_border(ws, start, 1, activitati_end_row, last_activitati_col, thickness="thin")

        # ------------- DAILY TABLE -------------
        r0 = 23

        ws.cell(r0,1,"DATA").font=bold
        ws.merge_cells(start_row=r0,start_column=1,end_row=r0+2,end_column=1)
        ws.cell(row=r0, column=1).alignment = center_wrapped

        ws.cell(r0,2,"ACTIVITATI").font=bold
        ws.merge_cells(start_row=r0,start_column=2,end_row=r0,end_column=2*nr_Activitati+1)
        ws.cell(row=r0, column=2).alignment = center_wrapped

        ws.cell(r0,2*nr_Activitati+2,"TOTAL").font=bold
        ws.merge_cells(start_row=r0,start_column=2*nr_Activitati+2,end_row=r0+2,end_column=2*nr_Activitati+2)
        ws.cell(row=r0, column=2*nr_Activitati+2).alignment = center_wrapped

        # headers A1-A4
        for i in range(len(activitati)):
            col = 2 + i*2

            cell = ws.cell(row=r0+1, column=col)
            cell.value = f"A{i+1}"
            cell.font = bold
            cell.alignment = center_wrapped

            ws.merge_cells(start_row=r0+1, start_column=col, end_row=r0+1, end_column=col+1)

            ws.cell(row=r0+2, column=col, value="nr ore").alignment=center
            ws.cell(row=r0+2, column=col+1, value="interval").alignment=center

        hours = defaultdict(float)
        durata = defaultdict(list)

        for e in entries:
            key = (e.activitate_id, e.date.day)
            hours[key] += e.nr_ore
            durata[key].append(str(e.durata))

        total_month = 0
        totals_activitate = {act.id: 0 for act in activitati}

        # days loop
        for d in range(1,days+1):

            row = r0+2+d
            ws.cell(row,1,f"{d} {month_name}").alignment=center
            ws.cell(row,1,f"{d} {month_name}").font=bold


            total_day = 0

            for act_id, i in activitate_index.items():

                col = 2+i*2

                h = hours.get((act_id, d), "")
                dur_list = durata.get((act_id, d), [])
                dur = ", ".join(dur_list) if dur_list else ""

                ws.cell(row,col,h).alignment=center
                ws.cell(row,col+1,dur).alignment=center

                if isinstance(h,(int,float)):
                    total_day += h
                    totals_activitate[act_id] += h

            ws.cell(row,nr_Activitati*2+2,total_day)
            ws.cell(row,nr_Activitati*2+2).alignment=center
            ws.cell(row,nr_Activitati*2+2).font=bold


            total_month += total_day

        # totals per lab
        end = r0+2+days+1
        ws.cell(end,1,"TOTAL").font=bold
        ws.cell(end,1,"TOTAL").alignment=center

        for act_id, i in activitate_index.items():
            col = 2+i*2
            ws.cell(end, col, totals_activitate[act_id])
            ws.cell(end, col).font=bold
            ws.cell(end, col).alignment=center


        ws.cell(end,nr_Activitati*2+2,total_month).font=bold
        ws.cell(end,nr_Activitati*2+2).alignment=center


        # borders
        apply_border(ws, r0, 1, end, nr_Activitati*2+2, thickness="thin")
                
        ws.cell(end+5,2,"Aprobat,")
        ws.cell(end+6,2,f"Responsabil laborator in cercetare-proiectare\n"
                        f"pentru activitatea {lab.name[-1]},")
        ws.merge_cells(start_row=end+6,start_column=2,end_row=end+6,end_column=3)
        ws.cell(end+7,2,f"{director.first_name} {director.last_name.upper()} ")
        ws.cell(end+8,2,f"Semnătura")

        ws.cell(end+5,8,"Aprobat,")
        ws.cell(end+6,8,f"Responsabil AUMOVIO pentru activitatea {lab.name[-1]},")
        ws.cell(end+7,8,(lab.responsabil_aumovio or "").strip())
        ws.merge_cells(start_row=end+6,start_column=8,end_row=end+6,end_column=9)
        ws.cell(end+8,8,f"Semnătura")
        for r in range(1, end+11):
            for c in range(1,nr_Activitati*2+4):
                if is_merged_cell(ws, r, c):
                    continue
                ws.cell(row=r, column=c).fill = wb_fill

        # auto‑size sheet
        autofit_sheet(ws)

        user_workbooks.append((user, wb))

    return user_workbooks

@require_http_methods(["GET"])
def export_excel(request):

    if not request.user.is_authenticated:
        return HttpResponseForbidden()

    profile = request.user.userprofile

    lab_id = request.GET.get("lab_id")
    month = int(request.GET.get("month"))
    year = int(request.GET.get("year"))

    try:
        lab_id_int = int(lab_id)
    except (TypeError, ValueError):
        return HttpResponseBadRequest("Invalid lab_id")

    lab = get_object_or_404(Lab, id=lab_id_int)

    is_admin = profile.role == "admin"
    if not is_admin:
        is_member = LabMembership.objects.filter(lab=lab, profile=profile).exists()
        if not is_member:
            return HttpResponseForbidden()

    # Every user exports only their own sheets.
    users = [request.user]

    director_membership = (
        LabMembership.objects.filter(lab=lab, role="director")
        .select_related("profile__user")
        .first()
    )
    director_user = (
        director_membership.profile.user
        if director_membership is not None
        else request.user
    )

    wb_upt = upt_workbook(lab,users,month,year,director_user)
    conti_user_workbooks= conti_workbook(lab, users, month, year, director_user)

    upt_buffer = BytesIO()

    wb_upt.save(upt_buffer)

    zip_buffer = BytesIO()

    with zipfile.ZipFile(zip_buffer,"w") as z:
        export_user = request.user
        z.writestr(
            f"L{lab.name[-1]} Pontaj {month} {year} {export_user.last_name.upper()} {export_user.first_name}.xlsx",
            upt_buffer.getvalue(),
        )
        for user, wb in conti_user_workbooks:
            user_buffer = BytesIO()
            wb.save(user_buffer)
            z.writestr(
                f"L{lab.name[-1]}_Pontaj_{month}_{months_to_RO(month)}_{user.last_name.upper()}_{user.first_name}.xlsx",
                user_buffer.getvalue(),
            )

    zip_buffer.seek(0)

    response = HttpResponse(zip_buffer,content_type="application/zip")
    response["Content-Disposition"] = f'attachment; filename="pontaj_{lab.name}_{month}_{year}.zip"'

    return response
