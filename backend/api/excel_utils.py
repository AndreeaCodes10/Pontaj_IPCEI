from openpyxl.styles import Border, Side

def apply_border(ws, start_row, start_col, end_row, end_col, thickness="thin"):
    """
    Apply borders to a rectangular region.
    thickness: 'thin' or 'bold'
    """

    style = "thin" if thickness == "thin" else "thick"
    side = Side(style=style)
    border = Border(left=side, right=side, top=side, bottom=side)

    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            ws.cell(row=r, column=c).border = border

def apply_table_border(ws, start_row, start_col, end_row, end_col):
    thin = Side(style="thin")
    thick = Side(style="thick")

    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):

            left = thick if c == start_col else thin
            right = thick if c == end_col else thin
            top = thick if r == start_row else thin
            bottom = thick if r == end_row else thin

            ws.cell(row=r, column=c).border = Border(
                left=left,
                right=right,
                top=top,
                bottom=bottom
            )

from openpyxl.styles import Alignment

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

def merge_center(ws, start_row, start_col, end_row, end_col, value=None):
    ws.merge_cells(
        start_row=start_row,
        start_column=start_col,
        end_row=end_row,
        end_column=end_col
    )

    cell = ws.cell(start_row, start_col)

    if value is not None:
        cell.value = value

    cell.alignment = CENTER

from openpyxl.styles import Font

BOLD = Font(bold=True)

def write_header(ws, row, col, value):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = BOLD
    cell.alignment = CENTER
    return cell